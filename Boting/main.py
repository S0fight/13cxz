from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import openpyxl
from openpyxl.styles import Font, Border, Side
import json
import os
from datetime import datetime
import telegram.error

# Токен бота от @BotFather
TOKEN = '8085809154:AAFhxq9Yqwh7_Sn__xU4p8knVxIro_35EfM'
EXCEL_FILE = 'shadow_data.xlsx'
JSON_FILE = 'shadow_data.json'

# Инициализация баз данных
def init_databases():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"
        headers = ["ID", "Номер", "Username", "Имя", "Фамилия", "Дата регистрации", "Последняя активность",
                  "Язык", "Тип чата", "Ссылка", "Фото", "Биография", "Участники", "Тип доступа", "Обновлено",
                  "Взаимные контакты", "Заблокирован", "Устройство", "Подписки", "Ограничения", "Изменение Username",
                  "Бот", "Премиум", "Последний вход", "Может加入群组", "Приватность", "Супергруппа", "Админ",
                  "Описание", "Дата создания", "Верификация"]
        ws.append(headers)

        for col in range(1, 32):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

        column_widths = {'A': 15, 'B': 20, 'C': 25, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 10, 'I': 15,
                        'J': 30, 'K': 10, 'L': 30, 'M': 15, 'N': 15, 'O': 20, 'P': 20, 'Q': 15, 'R': 20,
                        'S': 30, 'T': 20, 'U': 20, 'V': 10, 'W': 10, 'X': 20, 'Y': 15, 'Z': 20, 'AA': 15,
                        'AB': 10, 'AC': 30, 'AD': 20, 'AE': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(EXCEL_FILE)

    if not os.path.exists(JSON_FILE) or os.path.getsize(JSON_FILE) == 0:
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump([], f)
    else:
        try:
            with open(JSON_FILE, 'r', encoding='utf-8') as f:
                json.load(f)
        except json.JSONDecodeError:
            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump([], f)

# Проверка наличия пользователя в базе (JSON)
def check_user_json(user_id):
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for user in data:
            if str(user['id']) == str(user_id):
                return True
        return False
    except (json.JSONDecodeError, FileNotFoundError):
        return False

# Сохранение пользователя в Excel
def save_user_excel(user_id, phone, username, first_name, last_name, language_code, chat_type, link,
                    has_photo="нет", bio="нет", members="0", access_type="неизвестно", mutual_contacts="нет",
                    blocked="нет", device="неизвестно", subscriptions="нет", restrictions="нет",
                    username_change="нет", is_bot="нет", is_premium="нет", last_seen="неизвестно",
                    can_join_groups="неизвестно", privacy_settings="неизвестно", is_supergroup="нет",
                    is_admin="нет", description="нет", creation_date="неизвестно", verified="нет"):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    last_activity = reg_date
    updated = reg_date
    if not any(str(row[0].value) == str(user_id) for row in ws.iter_rows(min_row=2)):
        ws.append([user_id, phone, username or "нет", first_name, last_name or "скрыто", reg_date,
                  last_activity, language_code or "неизвестно", chat_type, link or "нет", has_photo,
                  bio, members, access_type, updated, mutual_contacts, blocked, device, subscriptions,
                  restrictions, username_change, is_bot, is_premium, last_seen, can_join_groups,
                  privacy_settings, is_supergroup, is_admin, description, creation_date, verified])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
        wb.save(EXCEL_FILE)
    wb.close()

# Сохранение пользователя в JSON
def save_user_json(user_id, phone, username, first_name, last_name, language_code, chat_type, link,
                   has_photo="нет", bio="нет", members="0", access_type="неизвестно", mutual_contacts="нет",
                   blocked="нет", device="неизвестно", subscriptions="нет", restrictions="нет",
                   username_change="нет", is_bot="нет", is_premium="нет", last_seen="неизвестно",
                   can_join_groups="неизвестно", privacy_settings="неизвестно", is_supergroup="нет",
                   is_admin="нет", description="нет", creation_date="неизвестно", verified="нет"):
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        data = []

    reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    last_activity = reg_date
    updated = reg_date
    if not check_user_json(user_id):
        user_data = {
            "id": user_id,
            "phone": phone,
            "username": username or "нет",
            "first_name": first_name,
            "last_name": last_name or "скрыто",
            "reg_date": reg_date,
            "last_activity": last_activity,
            "language": language_code or "неизвестно",
            "chat_type": chat_type,
            "profile_link": link or "нет",
            "has_photo": has_photo,
            "bio": bio,
            "members": members,
            "access_type": access_type,
            "updated": updated,
            "mutual_contacts": mutual_contacts,
            "blocked": blocked,
            "device": device,
            "subscriptions": subscriptions,
            "restrictions": restrictions,
            "username_change": username_change,
            "is_bot": is_bot,
            "is_premium": is_premium,
            "last_seen": last_seen,
            "can_join_groups": can_join_groups,
            "privacy_settings": privacy_settings,
            "is_supergroup": is_supergroup,
            "is_admin": is_admin,
            "description": description,
            "creation_date": creation_date,
            "verified": verified
        }
        data.append(user_data)
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

# Обновление данных пользователя в JSON
async def update_user_data(user_id, context):
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for user in data:
            if str(user['id']) == str(user_id):
                try:
                    chat = await context.bot.get_chat(user_id)
                    user['last_activity'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    user['updated'] = user['last_activity']
                    user['has_photo'] = "да" if chat.photo else "нет"
                    user['bio'] = chat.bio if chat.bio else "нет"
                    user['members'] = str(chat.member_count) if hasattr(chat, 'member_count') else "0"
                    user['access_type'] = "публичный" if chat.username else "приватный"
                    user['blocked'] = "да" if chat.has_restrictions else "нет"
                    user['restrictions'] = "да" if chat.restricted else "нет"
                    user['is_bot'] = "да" if chat.is_bot else "нет"
                    user['is_premium'] = "да" if chat.is_premium else "нет"
                    user['last_seen'] = chat.last_seen if hasattr(chat, 'last_seen') else "неизвестно"
                    user['can_join_groups'] = "да" if hasattr(chat, 'can_join_groups') and chat.can_join_groups else "нет"
                    user['privacy_settings'] = "ограничено" if chat.has_private_forwards else "открыто"
                    user['is_supergroup'] = "да" if chat.type == "supergroup" else "нет"
                    user['is_admin'] = "да" if chat.permissions and chat.permissions.can_change_info else "нет"
                    user['description'] = chat.description if chat.description else "нет"
                    user['creation_date'] = chat.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(chat, 'created_at') else "неизвестно"
                    user['verified'] = "да" if chat.verified else "нет"
                except telegram.error.TimedOut:
                    print(f"Таймаут при обновлении данных для ID {user_id}")
                except Exception as e:
                    print(f"Ошибка при обновлении данных для ID {user_id}: {e}")
                break
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except (json.JSONDecodeError, FileNotFoundError):
        pass

# Сохранение пользователя в обе базы
def save_user(user_id, phone, username, first_name, last_name, language_code, chat_type, link):
    save_user_excel(user_id, phone, username, first_name, last_name, language_code, chat_type, link)
    save_user_json(user_id, phone, username, first_name, last_name, language_code, chat_type, link)

# Команда /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    init_databases()
    chat_type = update.message.chat.type
    link = f"https://t.me/{user.username}" if user.username else "нет"

    if check_user_json(user.id):
        await update_user_data(user.id, context)
        try:
            await update.message.reply_text("Вы уже зарегистрированы!\n\n" + get_commands())
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /start")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
    else:
        welcome_text = (
            "🗂 Добро пожаловать, братишка!\n\n"
            "Подтверди согласие с [Политикой конфиденциальности](https://telegra.ph/Politika-v-otnoshenii-obrabotki-personalnyh-dannyh-01-10-3) "
            "и [Пользовательским соглашением](https://telegra.ph/Publichnaya-oferta-na-zaklyuchenie-licenzionnogo-dogovora-09-25).\n\n"
            "⚠️ Всё на твоей совести, делай по кайфу!"
        )
        keyboard = [[InlineKeyboardButton("Согласен, погнали!", callback_data='agree')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            await update.message.reply_text(welcome_text, reply_markup=reply_markup, parse_mode='Markdown')
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /start")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Список команд
def get_commands():
    return (
        "📋 Команды для шарящих:\n\n"
        "👤 Пробив по данным:\n"
        "• `/probiv 123456789` - полный разнос по ID\n"
        "• `/phone +79991234567` - копаем по номеру\n"
        "• `/username @example` - выцепляем по юзернейму\n\n"
        "ℹ️ Инфа о себе:\n"
        "• `/info` - всё, что есть на тебя\n\n"
        "📩 Кидай номер телефона, чтобы завести акк."
    )

# Обработка согласия
async def handle_agree(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user = query.from_user
    chat_type = query.message.chat.type
    link = f"https://t.me/{user.username}" if user.username else "нет"

    if check_user_json(user.id):
        await update_user_data(user.id, context)
        try:
            await query.edit_message_text("Ты уже в деле!\n\n" + get_commands())
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в handle_agree")
            await query.edit_message_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
    else:
        try:
            await query.edit_message_text(f"Здарова, {user.first_name}! Ща замутим регистрацию.")
            keyboard = [[KeyboardButton("Кинуть номерок", request_contact=True)]]
            reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
            await context.bot.send_message(chat_id=query.message.chat_id, text="Жми кнопку, кидай номер!", reply_markup=reply_markup)
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в handle_agree")
            await query.edit_message_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Обработка номера телефона
async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    user = update.message.from_user
    phone = contact.phone_number
    chat_type = update.message.chat.type
    link = f"https://t.me/{user.username}" if user.username else "нет"

    if check_user_json(user.id):
        await update_user_data(user.id, context)
        try:
            await update.message.reply_text("Ты уже в базе, братишка!\n\n" + get_commands())
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в handle_contact")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
    else:
        save_user(user.id, phone, user.username, user.first_name, user.last_name, user.language_code, chat_type, link)
        try:
            await update.message.reply_text(f"✅ Номер {phone} в деле! Добро пожаловать, пацан!\n\n" + get_commands())
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в handle_contact")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Команда /info
async def info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    await update_user_data(user.id, context)

    if not check_user_json(user.id):
        try:
            await update.message.reply_text("Сначала зарегайся, братишка!")
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /info")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for entry in data:
        if str(entry['id']) == str(user.id):
            info_text = (
                f"👤 Твоя инфа, пацан:\n"
                f"• ID: `{entry['id']}`\n"
                f"• Номер: {entry['phone']}\n"
                f"• Username: @{entry['username'] if entry['username'] != 'нет' else 'скрыто'}\n"
                f"• Имя: {entry['first_name']}\n"
                f"• Фамилия: {entry['last_name']}\n"
                f"• Дата регистрации: {entry['reg_date']}\n"
                f"• Последняя активность: {entry['last_activity']}\n"
                f"• Язык: {entry['language']}\n"
                f"• Тип чата: {entry['chat_type']}\n"
                f"• Ссылка на профиль: {entry['profile_link']}\n"
                f"• Фото профиля: {entry['has_photo']}\n"
                f"• Биография: {entry['bio']}\n"
                f"• Участники: {entry['members']}\n"
                f"• Тип доступа: {entry['access_type']}\n"
                f"• Дата обновления: {entry['updated']}\n"
                f"• Взаимные контакты: {entry['mutual_contacts']}\n"
                f"• Заблокирован: {entry['blocked']}\n"
                f"• Устройство: {entry['device']}\n"
                f"• Подписки: {entry['subscriptions']}\n"
                f"• Ограничения: {entry['restrictions']}\n"
                f"• Изменение Username: {entry['username_change']}\n"
                f"• Бот: {entry['is_bot']}\n"
                f"• Премиум: {entry['is_premium']}\n"
                f"• Последний вход: {entry['last_seen']}\n"
                f"• Может加入群组: {entry['can_join_groups']}\n"
                f"• Приватность: {entry['privacy_settings']}\n"
                f"• Супергруппа: {entry['is_supergroup']}\n"
                f"• Админ: {entry['is_admin']}\n"
                f"• Описание: {entry['description']}\n"
                f"• Дата создания: {entry['creation_date']}\n"
                f"• Верификация: {entry['verified']}"
            )
            try:
                await update.message.reply_text(info_text, parse_mode='Markdown')
            except telegram.error.TimedOut:
                print("Ошибка: Таймаут при отправке сообщения в /info")
                await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
            return

# Команда /probiv
async def probiv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    await update_user_data(user.id, context)

    if not check_user_json(user.id):
        try:
            await update.message.reply_text("Сначала зарегайся, братишка!")
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /probiv")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    args = context.args
    if not args:
        try:
            await update.message.reply_text("Кидай ID! Пример: `/probiv 123456789`", parse_mode='Markdown')
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /probiv")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    target_id = args[0]
    result = f"🔍 Пробив по ID {target_id} — держи полный расклад:\n"
    found_in_db = False

    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for entry in data:
        if str(entry['id']) == str(target_id):
            result += (
                f"• ID: `{entry['id']}`\n"
                f"• Номер: {entry['phone']}\n"
                f"• Username: @{entry['username'] if entry['username'] != 'нет' else 'скрыто'}\n"
                f"• Имя: {entry['first_name']}\n"
                f"• Фамилия: {entry['last_name']}\n"
                f"• Дата регистрации: {entry['reg_date']}\n"
                f"• Последняя активность: {entry['last_activity']}\n"
                f"• Язык: {entry['language']}\n"
                f"• Тип чата: {entry['chat_type']}\n"
                f"• Ссылка на профиль: {entry['profile_link']}\n"
                f"• Фото профиля: {entry['has_photo']}\n"
                f"• Биография: {entry['bio']}\n"
                f"• Участники: {entry['members']}\n"
                f"• Тип доступа: {entry['access_type']}\n"
                f"• Дата обновления: {entry['updated']}\n"
                f"• Взаимные контакты: {entry['mutual_contacts']}\n"
                f"• Заблокирован: {entry['blocked']}\n"
                f"• Устройство: {entry['device']}\n"
                f"• Подписки: {entry['subscriptions']}\n"
                f"• Ограничения: {entry['restrictions']}\n"
                f"• Изменение Username: {entry['username_change']}\n"
                f"• Бот: {entry['is_bot']}\n"
                f"• Премиум: {entry['is_premium']}\n"
                f"• Последний вход: {entry['last_seen']}\n"
                f"• Может加入群组: {entry['can_join_groups']}\n"
                f"• Приватность: {entry['privacy_settings']}\n"
                f"• Супергруппа: {entry['is_supergroup']}\n"
                f"• Админ: {entry['is_admin']}\n"
                f"• Описание: {entry['description']}\n"
                f"• Дата создания: {entry['creation_date']}\n"
                f"• Верификация: {entry['verified']}"
            )
            found_in_db = True
            break

    if not found_in_db:
        try:
            chat = await context.bot.get_chat(int(target_id))
            link = f"https://t.me/{chat.username}" if chat.username else "нет"
            has_photo = "да" if chat.photo else "нет"
            bio = chat.bio if chat.bio else "нет"
            members = str(chat.member_count) if hasattr(chat, 'member_count') else "0"
            access_type = "публичный" if chat.username else "приватный"
            blocked = "да" if chat.has_restrictions else "нет"
            restrictions = "да" if chat.restricted else "нет"
            is_bot = "да" if chat.is_bot else "нет"
            is_premium = "да" if chat.is_premium else "нет"
            last_seen = chat.last_seen if hasattr(chat, 'last_seen') else "неизвестно"
            can_join_groups = "да" if hasattr(chat, 'can_join_groups') and chat.can_join_groups else "нет"
            privacy_settings = "ограничено" if chat.has_private_forwards else "открыто"
            is_supergroup = "да" if chat.type == "supergroup" else "нет"
            is_admin = "да" if chat.permissions and chat.permissions.can_change_info else "нет"
            description = chat.description if chat.description else "нет"
            creation_date = chat.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(chat, 'created_at') else "неизвестно"
            verified = "да" if chat.verified else "нет"
            result += (
                f"• ID: `{chat.id}`\n"
                f"• Username: @{chat.username if chat.username else 'скрыто (анонимность)'}\n"
                f"• Имя: {chat.first_name if chat.first_name else 'скрыто (анонимность)'}\n"
                f"• Фамилия: {chat.last_name if chat.last_name else 'скрыто (анонимность)'}\n"
                f"• Тип: {chat.type}\n"
                f"• Ссылка на профиль: {link}\n"
                f"• Фото профиля: {has_photo}\n"
                f"• Биография: {bio}\n"
                f"• Участники: {members}\n"
                f"• Тип доступа: {access_type}\n"
                f"• Заблокирован: {blocked}\n"
                f"• Ограничения: {restrictions}\n"
                f"• Бот: {is_bot}\n"
                f"• Премиум: {is_premium}\n"
                f"• Последний вход: {last_seen}\n"
                f"• Может加入群组: {can_join_groups}\n"
                f"• Приватность: {privacy_settings}\n"
                f"• Супергруппа: {is_supergroup}\n"
                f"• Админ: {is_admin}\n"
                f"• Описание: {description}\n"
                f"• Дата создания: {creation_date}\n"
                f"• Верификация: {verified}"
            )
        except telegram.error.TimedOut:
            result += "❌ Таймаут при запросе данных из Telegram! Проверь инет, братишка!"
        except Exception:
            result += f"❌ ID {target_id} не найден или доступ закрыт!"

    try:
        await update.message.reply_text(result, parse_mode='Markdown')
    except telegram.error.TimedOut:
        print("Ошибка: Таймаут при отправке сообщения в /probiv")
        await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Команда /phone
async def probiv_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    await update_user_data(user.id, context)

    if not check_user_json(user.id):
        try:
            await update.message.reply_text("Сначала зарегайся, братишка!")
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /phone")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    args = context.args
    if not args:
        try:
            await update.message.reply_text("Кидай номер! Пример: `/phone +79991234567`", parse_mode='Markdown')
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /phone")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    target_phone = args[0]
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for entry in data:
        if entry['phone'] == target_phone:
            result = (
                f"📱 Пробив по номеру {target_phone} — лови инфу:\n"
                f"• ID: `{entry['id']}`\n"
                f"• Номер: {entry['phone']}\n"
                f"• Username: @{entry['username'] if entry['username'] != 'нет' else 'скрыто'}\n"
                f"• Имя: {entry['first_name']}\n"
                f"• Фамилия: {entry['last_name']}\n"
                f"• Дата регистрации: {entry['reg_date']}\n"
                f"• Последняя активность: {entry['last_activity']}\n"
                f"• Язык: {entry['language']}\n"
                f"• Тип чата: {entry['chat_type']}\n"
                f"• Ссылка на профиль: {entry['profile_link']}\n"
                f"• Фото профиля: {entry['has_photo']}\n"
                f"• Биография: {entry['bio']}\n"
                f"• Участники: {entry['members']}\n"
                f"• Тип доступа: {entry['access_type']}\n"
                f"• Дата обновления: {entry['updated']}\n"
                f"• Взаимные контакты: {entry['mutual_contacts']}\n"
                f"• Заблокирован: {entry['blocked']}\n"
                f"• Устройство: {entry['device']}\n"
                f"• Подписки: {entry['subscriptions']}\n"
                f"• Ограничения: {entry['restrictions']}\n"
                f"• Изменение Username: {entry['username_change']}\n"
                f"• Бот: {entry['is_bot']}\n"
                f"• Премиум: {entry['is_premium']}\n"
                f"• Последний вход: {entry['last_seen']}\n"
                f"• Может加入群组: {entry['can_join_groups']}\n"
                f"• Приватность: {entry['privacy_settings']}\n"
                f"• Супергруппа: {entry['is_supergroup']}\n"
                f"• Админ: {entry['is_admin']}\n"
                f"• Описание: {entry['description']}\n"
                f"• Дата создания: {entry['creation_date']}\n"
                f"• Верификация: {entry['verified']}"
            )
            try:
                await update.message.reply_text(result, parse_mode='Markdown')
            except telegram.error.TimedOut:
                print("Ошибка: Таймаут при отправке сообщения в /phone")
                await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
            return
    try:
        await update.message.reply_text(f"❌ Номер {target_phone} не найден в базе, братишка!")
    except telegram.error.TimedOut:
        print("Ошибка: Таймаут при отправке сообщения в /phone")
        await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Команда /username
async def probiv_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    await update_user_data(user.id, context)

    if not check_user_json(user.id):
        try:
            await update.message.reply_text("Сначала зарегайся, братишка!")
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /username")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    args = context.args
    if not args:
        try:
            await update.message.reply_text("Кидай юзернейм! Пример: `/username @example`", parse_mode='Markdown')
        except telegram.error.TimedOut:
            print("Ошибка: Таймаут при отправке сообщения в /username")
            await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
        return

    target_username = args[0].lstrip('@')
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for entry in data:
        if entry['username'] == target_username:
            result = (
                f"👤 Пробив по @{target_username} — лови всё, что есть:\n"
                f"• ID: `{entry['id']}`\n"
                f"• Номер: {entry['phone']}\n"
                f"• Username: @{entry['username'] if entry['username'] != 'нет' else 'скрыто'}\n"
                f"• Имя: {entry['first_name']}\n"
                f"• Фамилия: {entry['last_name']}\n"
                f"• Дата регистрации: {entry['reg_date']}\n"
                f"• Последняя активность: {entry['last_activity']}\n"
                f"• Язык: {entry['language']}\n"
                f"• Тип чата: {entry['chat_type']}\n"
                f"• Ссылка на профиль: {entry['profile_link']}\n"
                f"• Фото профиля: {entry['has_photo']}\n"
                f"• Биография: {entry['bio']}\n"
                f"• Участники: {entry['members']}\n"
                f"• Тип доступа: {entry['access_type']}\n"
                f"• Дата обновления: {entry['updated']}\n"
                f"• Взаимные контакты: {entry['mutual_contacts']}\n"
                f"• Заблокирован: {entry['blocked']}\n"
                f"• Устройство: {entry['device']}\n"
                f"• Подписки: {entry['subscriptions']}\n"
                f"• Ограничения: {entry['restrictions']}\n"
                f"• Изменение Username: {entry['username_change']}\n"
                f"• Бот: {entry['is_bot']}\n"
                f"• Премиум: {entry['is_premium']}\n"
                f"• Последний вход: {entry['last_seen']}\n"
                f"• Может加入群组: {entry['can_join_groups']}\n"
                f"• Приватность: {entry['privacy_settings']}\n"
                f"• Супергруппа: {entry['is_supergroup']}\n"
                f"• Админ: {entry['is_admin']}\n"
                f"• Описание: {entry['description']}\n"
                f"• Дата создания: {entry['creation_date']}\n"
                f"• Верификация: {entry['verified']}"
            )
            try:
                await update.message.reply_text(result, parse_mode='Markdown')
            except telegram.error.TimedOut:
                print("Ошибка: Таймаут при отправке сообщения в /username")
                await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")
            return
    try:
        await update.message.reply_text(f"❌ Юзернейм @{target_username} не найден в базе, братишка!")
    except telegram.error.TimedOut:
        print("Ошибка: Таймаут при отправке сообщения в /username")
        await update.message.reply_text("Ошибка: Не удалось подключиться к Telegram. Проверьте интернет и попробуйте снова.")

# Запуск бота
def main():
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_agree, pattern='agree'))
    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    app.add_handler(CommandHandler("info", info))
    app.add_handler(CommandHandler("probiv", probiv))
    app.add_handler(CommandHandler("phone", probiv_phone))
    app.add_handler(CommandHandler("username", probiv_username))

    print("✅ Бот запущен, всё пиздец как круто!")
    app.run_polling()

if __name__ == '__main__':
    main()